import openpyxl


def getRowCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column


def getCellData(path, sheetname, rowNum, colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowNum, column=colNum).value


def setCellData(path, sheetname, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)


path = "..//excel//testdata.xlsx"
sheetName = "LoginTest"
rows = getRowCount(path, sheetName)
cols = getColCount(path, sheetName)

print(rows, '---- ', cols)

print(getCellData(path, sheetName, 2, 1))
setCellData(path, sheetName, 1, 4, "DOB")
